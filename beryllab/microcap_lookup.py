import yfinance as yf
import json 
import openpyxl
import time 

# ticker = yf.Ticker('PARAA')
# formated_ticker_json = json.dumps(ticker.info, indent=4) # Use indent for pretty printing
# print("JSON String:")
# print(formated_ticker_json)

def save_workbook(workbook, file_path):
    """
    Save the workbook and handle potential errors
    """
    try:
        workbook.save(file_path)
        print("File saved successfully!")
        return True
    except PermissionError:
        print("Error: Cannot save the file. Please make sure the Excel file is closed and you have write permissions.")
        return False
    except Exception as e:
        print(f"Error saving file: {e}")
        return False

# Load the workbook
workbook = openpyxl.load_workbook('./Microcap_Stock_Screener.xlsx')
file_path = './Microcap_Stock_Screener.xlsx'


# Select a sheet
sheet_screener = workbook['MicroCap']  # Get the main sheet
sheet_contacts = workbook['Contacts'] # Get the contacts sheet

# Empty the Contacts sheet except for the header row
while sheet_contacts.max_row > 1:  # Preserve header row
    sheet_contacts.delete_rows(2)  # Always delete row 2 until only header remains
print("Contacts sheet cleared, ready for new data")

maxrow = 1246 # the last record row of the sheet
start_row = 2 # the first record row of the sheet, chnge this if you want to start from 

#Iterate through rows
records_processed = 0
for rowNum in range(start_row, maxrow + 1):  # Added +1 to include the last row
    try:
           # Add delay between requests
        time.sleep(1)  # 1 second delay
        # Get ticker symbol from column 2 and clean it
        ticker_symbol_raw = str(sheet_screener.cell(row=rowNum, column=2).value).strip()
        company_name = str(sheet_screener.cell(row=rowNum, column=3).value)
        if not ticker_symbol_raw:  # Skip empty rows
            continue
            
        # Format ticker symbol: replace / with - and ensure uppercase
        ticker_symbol = ticker_symbol_raw.replace('/', '-').upper()
        print(f"Processing ticker: {ticker_symbol}")  # Added for progress tracking
        
        # Get ticker info
        try:
            ticker = yf.Ticker(ticker_symbol)
            info = ticker.info 
        except Exception as e:
            print(f"Error fetching data for {ticker_symbol}: {e}")
            continue

        # Get the info from the ticker
        website = info.get("website", "N/A")  # Use N/A if website not found
        longBusinessSummary = info.get("longBusinessSummary", "N/A")
        fullTimeEmployees= info.get("fullTimeEmployees", "N/A")
        address1 = info.get("address1", "N/A")
        city = info.get("city", "N/A")
        state = info.get("state", "N/A")
        country = info.get("country", "N/A")
        phone = info.get("phone", "N/A")
        averageVolume = info.get("averageVolume", "N/A")
        companyOfficers = info.get("companyOfficers", None)
        date = time.strftime("%Y-%m-%d %H:%M:%S")

        # Update website in columns needed  
        sheet_screener.cell(row=rowNum, column=1).value = date
        sheet_screener.cell(row=rowNum, column=6).value = website
        sheet_screener.cell(row=rowNum, column=8).value = longBusinessSummary
        sheet_screener.cell(row=rowNum, column=10).value = fullTimeEmployees
        sheet_screener.cell(row=rowNum, column=11).value = address1
        sheet_screener.cell(row=rowNum, column=12).value = city
        sheet_screener.cell(row=rowNum, column=13).value = state
        sheet_screener.cell(row=rowNum, column=14).value = country
        sheet_screener.cell(row=rowNum, column=15).value = phone
        sheet_screener.cell(row=rowNum, column=18).value = averageVolume
        # Check if companyOfficers exists and is not empty
        if companyOfficers is not None and len(companyOfficers) > 0:
            for person in companyOfficers:
                row_data = [
                    date,
                    ticker_symbol_raw, 
                    company_name,
                    person.get("name", "N/A"),
                    person.get("title", "N/A"), 
                    person.get("yearBorn", "N/A")
                ]
                sheet_contacts.append(row_data)
            
        records_processed += 1
        # Save every 100 records
        if records_processed % 100 == 0:
            if not save_workbook(workbook, file_path):
                break  # Stop processing if save fails
    except Exception as e:
        print(f"Error processing row {rowNum}, ticker: {ticker_symbol}. Error: {e}")
        save_workbook(workbook, file_path)
        break

# Final save for remaining records
if records_processed % 100 != 0:
    save_workbook(workbook, file_path)

workbook.close()
print("Screener updated successfully!")